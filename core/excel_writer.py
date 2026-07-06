# -*- coding: utf-8 -*-
"""
Excel输出模块

功能：
  1. Sheet "最新价格" — 每次运行覆盖，展示当前快照
  2. Sheet "价格历史" — 每次运行追加，记录所有历史价格
  3. Sheet "价格统计" — 使用公式自动计算最低/最高/平均价

Excel中图片列存储本地文件路径，点击可直接打开查看。
"""

import logging
import statistics
import math
from pathlib import Path
from datetime import datetime
from collections import Counter

import numpy as np

from openpyxl import Workbook, load_workbook

from .config_loader import get_pricing_config
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 样式常量
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="4472C4")
CELL_FONT = Font(name="微软雅黑", size=10)
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 列定义
LATEST_COLUMNS = [
    ("抓取时间", 20),
    ("游戏名", 22),
    ("版本", 8),
    ("商品标题", 40),
    ("价格", 10),
    ("卖家", 14),
    ("地区", 12),
    ("商品链接", 30),
    ("宣传图路径", 40),
    ("卡带信息图路径", 45),
    ("商品描述", 50),
    ("成色", 12),
]

HISTORY_COLUMNS = [
    ("抓取日期", 12),
    ("抓取时间", 20),
    ("游戏名", 22),
    ("版本", 8),
    ("商品ID", 14),
    ("商品标题", 40),
    ("价格", 10),
    ("卖家", 14),
    ("地区", 12),
    ("商品链接", 30),
]


class ExcelWriter:
    """Excel写入器：管理多Sheet工作簿"""

    def __init__(self, output_path: str):
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def write(self, all_items: list) -> str:
        """
        写入或更新Excel文件

        如果文件已存在：覆盖"最新价格"sheet，追加"价格历史"sheet
        如果文件不存在：创建新文件，包含所有sheet
        """
        if self.output_path.exists():
            wb = load_workbook(self.output_path)
            self._update_latest_sheet(wb, all_items)
            self._append_history_sheet(wb, all_items)
            self._update_stats_sheet(wb)
            self._update_summary_sheet(wb, all_items)
        else:
            wb = Workbook()
            self._create_latest_sheet(wb, all_items)
            self._create_history_sheet(wb, all_items)
            self._create_stats_sheet(wb)
            self._create_summary_sheet(wb, all_items)

        wb.save(self.output_path)
        logger.info("Excel已保存: %s", self.output_path)
        return str(self.output_path)

    # ------------------------------------------------------------------
    # 最新价格 Sheet
    # ------------------------------------------------------------------

    def _create_latest_sheet(self, wb: Workbook, items: list):
        """创建"最新价格"sheet"""
        ws = wb.active
        ws.title = "最新价格"
        self._write_headers(ws, LATEST_COLUMNS)
        self._write_items_latest(ws, items, start_row=2)
        self._freeze_header(ws)

    def _update_latest_sheet(self, wb, items: list):
        """更新"最新价格"sheet（清空旧数据，写入新数据）"""
        if "最新价格" in wb.sheetnames:
            ws = wb["最新价格"]
            # 清空旧数据（保留表头）
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
        else:
            ws = wb.create_sheet("最新价格")
            self._write_headers(ws, LATEST_COLUMNS)

        self._write_items_latest(ws, items, start_row=2)
        self._freeze_header(ws)

    def _write_items_latest(self, ws, items: list, start_row: int):
        """写入商品数据到最新价格sheet"""
        for i, item in enumerate(items):
            row = start_row + i
            promo_imgs = item.get("downloaded_images", {}).get("promo_images", [])
            detail_imgs = item.get("downloaded_images", {}).get("detail_images", [])

            values = [
                item.get("scrape_date", ""),
                item.get("game_name", ""),
                item.get("version", ""),
                item.get("title", ""),
                item.get("price", ""),
                item.get("seller", ""),
                item.get("area", ""),
                item.get("url", ""),
                "\n".join(promo_imgs),
                "\n".join(detail_imgs),
                (item.get("description") or "")[:200],
                item.get("condition", ""),
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = CELL_FONT
                cell.alignment = LEFT_WRAP
                cell.border = THIN_BORDER
                if i % 2 == 1:
                    cell.fill = ALT_FILL

            # 价格列数字格式
            price_cell = ws.cell(row=row, column=5)
            if isinstance(price_cell.value, (int, float)):
                price_cell.number_format = "¥#,##0.00"
                price_cell.alignment = CENTER

    # ------------------------------------------------------------------
    # 价格历史 Sheet
    # ------------------------------------------------------------------

    def _create_history_sheet(self, wb: Workbook, items: list):
        """创建"价格历史"sheet"""
        ws = wb.create_sheet("价格历史")
        self._write_headers(ws, HISTORY_COLUMNS)
        self._write_items_history(ws, items, start_row=2)
        self._freeze_header(ws)

    def _append_history_sheet(self, wb, items: list):
        """追加数据到"价格历史"sheet"""
        if "价格历史" in wb.sheetnames:
            ws = wb["价格历史"]
        else:
            ws = wb.create_sheet("价格历史")
            self._write_headers(ws, HISTORY_COLUMNS)

        start_row = ws.max_row + 1
        self._write_items_history(ws, items, start_row=start_row)
        self._freeze_header(ws)

    def _write_items_history(self, ws, items: list, start_row: int):
        """写入历史记录"""
        for i, item in enumerate(items):
            row = start_row + i
            scrape_date = item.get("scrape_date", "")
            # 从 "2026-07-06 14:30:00" 提取日期部分
            date_only = scrape_date[:10] if len(scrape_date) >= 10 else scrape_date

            values = [
                date_only,
                scrape_date,
                item.get("game_name", ""),
                item.get("version", ""),
                item.get("itemId", ""),
                item.get("title", ""),
                item.get("price", ""),
                item.get("seller", ""),
                item.get("area", ""),
                item.get("url", ""),
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = CELL_FONT
                cell.alignment = LEFT_WRAP
                cell.border = THIN_BORDER

            # 价格列数字格式
            price_cell = ws.cell(row=row, column=7)
            if isinstance(price_cell.value, (int, float)):
                price_cell.number_format = "¥#,##0.00"
                price_cell.alignment = CENTER

    # ------------------------------------------------------------------
    # 价格统计 Sheet（使用Excel公式，不硬编码计算结果）
    # ------------------------------------------------------------------

    def _create_stats_sheet(self, wb: Workbook):
        """创建"价格统计"sheet，使用公式自动计算"""
        ws = wb.create_sheet("价格统计")

        headers = [
            ("游戏名", 22),
            ("版本", 8),
            ("最低价", 12),
            ("最高价", 12),
            ("平均价", 12),
            ("记录数", 10),
            ("最近抓取日期", 20),
        ]
        self._write_headers(ws, headers)

        # 获取所有游戏名（从价格历史sheet推断）
        # 首次创建时，统计sheet只有表头，公式在_update_stats_sheet中填充
        self._freeze_header(ws)

    def _update_stats_sheet(self, wb):
        """更新统计sheet，为每个游戏写入公式行"""
        if "价格统计" not in wb.sheetnames:
            ws = wb.create_sheet("价格统计")
            headers = [
                ("游戏名", 22),
                ("版本", 8),
                ("最低价", 12),
                ("最高价", 12),
                ("平均价", 12),
                ("记录数", 10),
                ("最近抓取日期", 20),
            ]
            self._write_headers(ws, headers)
        else:
            ws = wb["价格统计"]

        # 清空旧数据
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # 从价格历史sheet获取游戏列表
        if "价格历史" not in wb.sheetnames:
            return

        hist_ws = wb["价格历史"]
        hist_last_row = hist_ws.max_row

        # 收集所有游戏名（去重，保持顺序）
        games = []
        seen = set()
        for row in range(2, hist_last_row + 1):
            name = hist_ws.cell(row=row, column=3).value
            version = hist_ws.cell(row=row, column=4).value or ""
            if name and name not in seen:
                games.append((name, version))
                seen.add(name)

        # 为每个游戏写入统计公式行
        # 价格历史sheet中：C列=游戏名(3), G列=价格(7), A列=抓取日期(1)
        for i, (game_name, version) in enumerate(games):
            row = i + 2
            ws.cell(row=row, column=1, value=game_name).font = CELL_FONT
            ws.cell(row=row, column=2, value=version).font = CELL_FONT

            # 最低价 — 使用MINIFS（Excel 2019+ / LibreOffice支持）
            ws.cell(
                row=row, column=3,
                value=f'=MINIFS(价格历史!G:G,价格历史!C:C,A{row})',
            )
            # 最高价
            ws.cell(
                row=row, column=4,
                value=f'=MAXIFS(价格历史!G:G,价格历史!C:C,A{row})',
            )
            # 平均价
            ws.cell(
                row=row, column=5,
                value=f'=AVERAGEIF(价格历史!C:C,A{row},价格历史!G:G)',
            )
            # 记录数
            ws.cell(
                row=row, column=6,
                value=f'=COUNTIF(价格历史!C:C,A{row})',
            )
            # 最近抓取日期
            ws.cell(
                row=row, column=7,
                value=f'=MAXIFS(价格历史!A:A,价格历史!C:C,A{row})',
            )

            # 设置格式
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = CELL_FONT
                cell.alignment = CENTER
                cell.border = THIN_BORDER
                if i % 2 == 1:
                    cell.fill = ALT_FILL

            # 价格列格式
            for col in [3, 4, 5]:
                ws.cell(row=row, column=col).number_format = "¥#,##0.00"

        self._freeze_header(ws)

    # ------------------------------------------------------------------
    # 游戏价格汇总 Sheet（每游戏一行，价格取中位数）
    # ------------------------------------------------------------------

    SUMMARY_COLUMNS = [
        ("游戏名", 22),
        ("版本", 8),
        ("样本数", 10),
        ("参考价", 12),
        ("最低价", 12),
        ("最高价", 12),
        ("平均价", 12),
        ("抓取日期", 20),
    ]

    def _create_summary_sheet(self, wb: Workbook, items: list):
        """创建"游戏价格汇总"sheet"""
        ws = wb.create_sheet("游戏价格汇总")
        self._write_headers(ws, self.SUMMARY_COLUMNS)
        self._write_summary_rows(ws, items, start_row=2)
        self._freeze_header(ws)

    def _update_summary_sheet(self, wb, items: list):
        """更新"游戏价格汇总"sheet（覆盖旧数据）"""
        if "游戏价格汇总" in wb.sheetnames:
            ws = wb["游戏价格汇总"]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
        else:
            ws = wb.create_sheet("游戏价格汇总")
            self._write_headers(ws, self.SUMMARY_COLUMNS)

        self._write_summary_rows(ws, items, start_row=2)
        self._freeze_header(ws)

    @staticmethod
    def _estimate_price(prices: list) -> float:
        """
        二手卡带推荐价格算法

        流程：
            1. IQR 去异常值
            2. 价格分桶（5元）
            3. 找最大价格簇（市场主流价格）
            4. 主价格簇内取中位数
            5. 与整体中位数融合
            6. 回归市场习惯价（5元取整）
        """
        if not prices:
            return 0

        arr = sorted(float(p) for p in prices)
        n = len(arr)

        # 样本太少直接中位数
        if n < 10:
            return round(statistics.median(arr), 2)

        # ---------- IQR 去异常 ----------
        q1 = np.percentile(arr, 25)
        q3 = np.percentile(arr, 75)

        iqr = q3 - q1

        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr

        cleaned = [
            p for p in arr
            if lower <= p <= upper
        ]

        if len(cleaned) < 5:
            return round(statistics.median(arr), 2)

        # ---------- 价格分桶 ----------
        # 从配置读取分桶大小
        bucket_size = get_pricing_config().get("bucket_size", 10)

        buckets = Counter(
            round(p / bucket_size) * bucket_size
            for p in cleaned
        )

        # 找主价格簇
        main_bucket, count = buckets.most_common(1)[0]

        # ---------- 主价格区 ----------
        hot_zone = [
            p for p in cleaned
            if abs(p - main_bucket) <= bucket_size
        ]

        # 主价格区中位数
        hot_price = (
            statistics.median(hot_zone)
            if hot_zone
            else main_bucket
        )

        # 整体中位数
        median_price = statistics.median(cleaned)

        # ---------- 融合 ----------
        # 卡带市场：主流成交价权重大一些
        final_price = (
            hot_price * 0.7 +
            median_price * 0.3
        )

        # ---------- 稳定性约束 ----------
        # 最终价不能偏离中位数超过 30%，防止小样本抖动
        lower_bound = median_price * 0.7
        upper_bound = median_price * 1.3
        final_price = max(lower_bound, min(final_price, upper_bound))

        # ---------- 1元向上取整 ----------
        final_price = math.ceil(final_price)

        return float(final_price)

    def _write_summary_rows(self, ws, items: list, start_row: int):
        """按游戏分组，计算统计值写入汇总sheet（使用 IQR+KDE+四分位策略）"""
        # 按游戏名分组
        game_groups = {}
        for item in items:
            name = item.get("game_name", "")
            if not name:
                continue
            if name not in game_groups:
                game_groups[name] = {"version": item.get("version", ""), "prices": []}
            price = item.get("price")
            if isinstance(price, (int, float)) and price > 0:
                game_groups[name]["prices"].append(price)

        scrape_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for i, (game_name, info) in enumerate(sorted(game_groups.items())):
            row = start_row + i
            prices = info["prices"]
            count = len(prices)

            ws.cell(row=row, column=1, value=game_name).font = CELL_FONT
            ws.cell(row=row, column=2, value=info["version"]).font = CELL_FONT
            ws.cell(row=row, column=3, value=count).font = CELL_FONT

            if count > 0:
                final_price = self._estimate_price(prices)
                min_price = min(prices)
                max_price = max(prices)
                avg_price = sum(prices) / count

                ws.cell(row=row, column=4, value=final_price)
                ws.cell(row=row, column=5, value=round(min_price, 2))
                ws.cell(row=row, column=6, value=round(max_price, 2))
                ws.cell(row=row, column=7, value=round(avg_price, 2))
            else:
                ws.cell(row=row, column=4, value="")
                ws.cell(row=row, column=5, value="")
                ws.cell(row=row, column=6, value="")
                ws.cell(row=row, column=7, value="")

            ws.cell(row=row, column=8, value=scrape_date).font = CELL_FONT

            # 格式
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.font = CELL_FONT
                cell.alignment = CENTER
                cell.border = THIN_BORDER
                if i % 2 == 1:
                    cell.fill = ALT_FILL

            # 价格列数字格式
            for col in [4, 5, 6, 7]:
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "¥#,##0.00"

    # ------------------------------------------------------------------
    # 通用工具
    # ------------------------------------------------------------------

    def _write_headers(self, ws, columns: list):
        """写入表头行"""
        for col, (title, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 28

    def _freeze_header(self, ws):
        """冻结首行"""
        ws.freeze_panes = "A2"

    def _auto_width(self, ws):
        """简单自动列宽"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    length = len(str(cell.value or ""))
                    if length > max_len:
                        max_len = length
                except Exception:
                    pass
            adjusted = min(max_len + 4, 50)
            if ws.column_dimensions[col_letter].width:
                ws.column_dimensions[col_letter].width = max(
                    ws.column_dimensions[col_letter].width, adjusted
                )
