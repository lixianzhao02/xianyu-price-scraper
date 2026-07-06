# -*- coding: utf-8 -*-
"""
生成示例Excel — 用模拟数据展示输出格式
"""
import sys
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, str(Path(__file__).parent))

from core.excel_writer import ExcelWriter

# 模拟数据：模拟两次抓取（今天和昨天）以展示价格历史
today = datetime.now()
yesterday = today - timedelta(days=1)

mock_items = [
    # === 塞尔达传说：王国之泪 ===
    {
        "game_name": "塞尔达传说：王国之泪",
        "version": "卡带",
        "itemId": "72831456",
        "title": "塞尔达传说 王国之泪 NS卡带 9成新 带原盒",
        "price": 185.0,
        "priceText": "¥185",
        "seller": "游戏达人小铺",
        "area": "上海",
        "url": "https://www.goofish.com/item/72831456",
        "image": "https://img.alicdn.com/imgextra/i3/example1.jpg",
        "detail_images": [
            "https://img.alicdn.com/imgextra/i3/example1_detail1.jpg",
            "https://img.alicdn.com/imgextra/i3/example1_detail2.jpg",
        ],
        "description": "塞尔达传说王国之泪卡带，9成新，带原盒说明书，通关出。卡带状态良好无划痕。",
        "condition": "9成新",
        "scrape_date": today.strftime("%Y-%m-%d %H:%M:%S"),
        "downloaded_images": {
            "promo_images": ["data/images/塞尔达传说_王国之泪/72831456/promo_1.jpg"],
            "detail_images": [
                "data/images/塞尔达传说_王国之泪/72831456/detail_1.jpg",
                "data/images/塞尔达传说_王国之泪/72831456/detail_2.jpg",
            ],
        },
    },
    {
        "game_name": "塞尔达传说：王国之泪",
        "version": "卡带",
        "itemId": "72831457",
        "title": "NS卡带 塞尔达王国之泪 95新 国行",
        "price": 175.0,
        "priceText": "¥175",
        "seller": "二手游戏回收",
        "area": "广东深圳",
        "url": "https://www.goofish.com/item/72831457",
        "image": "https://img.alicdn.com/imgextra/i4/example2.jpg",
        "detail_images": ["https://img.alicdn.com/imgextra/i4/example2_detail1.jpg"],
        "description": "国行塞尔达王国之泪，95新，仅通关一次，包装齐全。",
        "condition": "95新",
        "scrape_date": today.strftime("%Y-%m-%d %H:%M:%S"),
        "downloaded_images": {
            "promo_images": ["data/images/塞尔达传说_王国之泪/72831457/promo_1.jpg"],
            "detail_images": ["data/images/塞尔达传说_王国之泪/72831457/detail_1.jpg"],
        },
    },
    {
        "game_name": "塞尔达传说：王国之泪",
        "version": "卡带",
        "itemId": "72831458",
        "title": "【急出】塞尔达王国之泪卡带 switch",
        "price": 168.0,
        "priceText": "¥168",
        "seller": "闲置回血",
        "area": "北京",
        "url": "https://www.goofish.com/item/72831458",
        "image": "https://img.alicdn.com/imgextra/i5/example3.jpg",
        "detail_images": [],
        "description": "急出回血，塞尔达王国之泪卡带，8成新无盒。",
        "condition": "8成新",
        "scrape_date": today.strftime("%Y-%m-%d %H:%M:%S"),
        "downloaded_images": {
            "promo_images": ["data/images/塞尔达传说_王国之泪/72831458/promo_1.jpg"],
            "detail_images": [],
        },
    },
    # === 马力欧卡丁车8豪华版 ===
    {
        "game_name": "马力欧卡丁车8豪华版",
        "version": "卡带",
        "itemId": "72831460",
        "title": "马里奥赛车8豪华版 NS卡带 9成新带盒",
        "price": 145.0,
        "priceText": "¥145",
        "seller": "快乐游戏屋",
        "area": "浙江杭州",
        "url": "https://www.goofish.com/item/72831460",
        "image": "https://img.alicdn.com/imgextra/i6/example4.jpg",
        "detail_images": ["https://img.alicdn.com/imgextra/i6/example4_detail1.jpg"],
        "description": "马力欧卡丁车8豪华版卡带，9成新，带原盒，多人联机超好玩。",
        "condition": "9成新",
        "scrape_date": today.strftime("%Y-%m-%d %H:%M:%S"),
        "downloaded_images": {
            "promo_images": ["data/images/马力欧卡丁车8豪华版/72831460/promo_1.jpg"],
            "detail_images": ["data/images/马力欧卡丁车8豪华版/72831460/detail_1.jpg"],
        },
    },
    {
        "game_name": "马力欧卡丁车8豪华版",
        "version": "卡带",
        "itemId": "72831461",
        "title": "Switch 马里奥赛车8 卡带 国行全新未拆",
        "price": 199.0,
        "priceText": "¥199",
        "seller": "正品数码专营",
        "area": "广东广州",
        "url": "https://www.goofish.com/item/72831461",
        "image": "https://img.alicdn.com/imgextra/i7/example5.jpg",
        "detail_images": ["https://img.alicdn.com/imgextra/i7/example5_detail1.jpg"],
        "description": "国行全新未拆封，马力欧卡丁车8豪华版，包邮。",
        "condition": "全新",
        "scrape_date": today.strftime("%Y-%m-%d %H:%M:%S"),
        "downloaded_images": {
            "promo_images": ["data/images/马力欧卡丁车8豪华版/72831461/promo_1.jpg"],
            "detail_images": ["data/images/马力欧卡丁车8豪华版/72831461/detail_1.jpg"],
        },
    },
    # === 动物森友会 ===
    {
        "game_name": "集合啦！动物森友会",
        "version": "卡带",
        "itemId": "72831465",
        "title": "动物森友会 NS卡带 switch卡带 95新",
        "price": 138.0,
        "priceText": "¥138",
        "seller": "岛主退坑出",
        "area": "四川成都",
        "url": "https://www.goofish.com/item/72831465",
        "image": "https://img.alicdn.com/imgextra/i8/example6.jpg",
        "detail_images": ["https://img.alicdn.com/imgextra/i8/example6_detail1.jpg"],
        "description": "动森卡带，95新带盒，岛上装修已完成，开心出给有缘人。",
        "condition": "95新",
        "scrape_date": today.strftime("%Y-%m-%d %H:%M:%S"),
        "downloaded_images": {
            "promo_images": ["data/images/集合啦_动物森友会/72831465/promo_1.jpg"],
            "detail_images": ["data/images/集合啦_动物森友会/72831465/detail_1.jpg"],
        },
    },
    {
        "game_name": "集合啦！动物森友会",
        "version": "卡带",
        "itemId": "72831466",
        "title": "【包邮】动物森友会 卡带 9成新",
        "price": 132.0,
        "priceText": "¥132",
        "seller": "闲鱼玩家",
        "area": "江苏南京",
        "url": "https://www.goofish.com/item/72831466",
        "image": "https://img.alicdn.com/imgextra/i9/example7.jpg",
        "detail_images": [],
        "description": "动森卡带出，9成新，无原盒，包邮。",
        "condition": "9成新",
        "scrape_date": today.strftime("%Y-%m-%d %H:%M:%S"),
        "downloaded_images": {
            "promo_images": ["data/images/集合啦_动物森友会/72831466/promo_1.jpg"],
            "detail_images": [],
        },
    },
    # === 宝可梦 朱 ===
    {
        "game_name": "宝可梦 朱",
        "version": "卡带",
        "itemId": "72831470",
        "title": "宝可梦朱 NS卡带 通关出 9成新",
        "price": 155.0,
        "priceText": "¥155",
        "seller": "宝可梦训练师",
        "area": "湖北武汉",
        "url": "https://www.goofish.com/item/72831470",
        "image": "https://img.alicdn.com/imgextra/i1/example8.jpg",
        "detail_images": ["https://img.alicdn.com/imgextra/i1/example8_detail1.jpg"],
        "description": "宝可梦朱卡带，通关出，9成新带盒，存档可清空。",
        "condition": "9成新",
        "scrape_date": today.strftime("%Y-%m-%d %H:%M:%S"),
        "downloaded_images": {
            "promo_images": ["data/images/宝可梦_朱/72831470/promo_1.jpg"],
            "detail_images": ["data/images/宝可梦_朱/72831470/detail_1.jpg"],
        },
    },
    # === 斯普拉遁3 ===
    {
        "game_name": "斯普拉遁3 (Splatoon 3)",
        "version": "卡带",
        "itemId": "72831475",
        "title": "喷射战士3 Splatoon3 NS卡带 全新",
        "price": 210.0,
        "priceText": "¥210",
        "seller": "日版游戏专营",
        "area": "上海",
        "url": "https://www.goofish.com/item/72831475",
        "image": "https://img.alicdn.com/imgextra/i2/example9.jpg",
        "detail_images": ["https://img.alicdn.com/imgextra/i2/example9_detail1.jpg"],
        "description": "喷射战士3日版卡带，全新未拆，支持中文。",
        "condition": "全新",
        "scrape_date": today.strftime("%Y-%m-%d %H:%M:%S"),
        "downloaded_images": {
            "promo_images": ["data/images/斯普拉遁3_Splatoon_3/72831475/promo_1.jpg"],
            "detail_images": ["data/images/斯普拉遁3_Splatoon_3/72831475/detail_1.jpg"],
        },
    },
]

# 昨天的历史数据（模拟价格变化）
history_items = []
for item in mock_items:
    hist = dict(item)
    hist["price"] = item["price"] + 5  # 昨天贵5块
    hist["priceText"] = f"¥{hist['price']}"
    hist["scrape_date"] = yesterday.strftime("%Y-%m-%d %H:%M:%S")
    history_items.append(hist)

# 先写入昨天的数据（作为历史）
output_path = Path(__file__).parent / "data" / "output" / "NS卡带价格追踪.xlsx"
output_path.parent.mkdir(parents=True, exist_ok=True)

writer = ExcelWriter(str(output_path))

# 第一次写入：昨天的数据（创建文件+所有sheet）
print("写入昨日历史数据...")
writer.write(history_items)

# 第二次写入：今天的数据（追加历史+更新最新价格）
print("写入今日最新数据...")
writer.write(mock_items)

print(f"\n示例Excel已生成: {output_path}")
print(f"文件大小: {output_path.stat().st_size / 1024:.1f} KB")

# 读取并展示各Sheet内容
from openpyxl import load_workbook
wb = load_workbook(str(output_path))
print(f"\nSheet列表: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name}  (共{ws.max_row}行 x {ws.max_column}列)")
    print(f"{'='*80}")
    
    # 打印表头
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    print("  表头:", " | ".join(str(h) for h in headers if h))
    
    # 打印前5行数据
    for row in range(2, min(ws.max_row+1, 7)):
        vals = []
        for col in range(1, ws.max_column+1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                vals.append("")
            elif isinstance(v, str) and len(v) > 25:
                vals.append(v[:25] + "...")
            else:
                vals.append(str(v))
        print(f"  行{row}: {' | '.join(vals)}")
    
    if ws.max_row > 6:
        print(f"  ... (共{ws.max_row-1}条数据)")
